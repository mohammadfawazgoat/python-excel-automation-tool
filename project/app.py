import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from pathlib import Path

def main():
    filename = input("Enter filename along with extention:")
    path = Path(filename)
    if not path.exists():
        print("File does not exist.")
        return 
    updated_value = float(input("Enter the value to be multiplied with:"))
    updated_filename = input("Enter the name of the updated file along with extention:")
    path = Path(updated_filename)
    if path.exists():
        print("File already exists.")
        return
    process_workbook(filename, updated_value,updated_filename)
    

def process_workbook(filename, updated_valuee,updated_filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        updated_value = cell.value * updated_valuee
        updated_cell = sheet.cell(row, 4)
        updated_cell.value = updated_value

    value = Reference(sheet,
            min_row = 2,
            max_row = sheet.max_row,
            min_col = 4,
            max_col = 4)
    chart = BarChart()
    chart.add_data(value)
    sheet.add_chart(chart,"A7")
    wb.save(updated_filename)    


if __name__ == "__main__":
    main()    